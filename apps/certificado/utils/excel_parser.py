"""
Parser de archivos Excel con estudiantes.

Extrae y valida los datos de estudiantes desde archivos Excel.
"""

import openpyxl
import re
import logging
from typing import List, Dict
from django.core.exceptions import ValidationError


logger = logging.getLogger(__name__)


class ExcelParseError(Exception):
    """
    Error personalizado para errores de parsing de Excel.
    """
    pass


class ExcelParser:
    """
    Clase para parsear archivos Excel con estudiantes.
    
    Formato esperado:
        - Columna 1: NOMBRES COMPLETOS
        - Columna 2: CORREO ELECTRONICO
    
    Retorna lista de diccionarios con los datos validados.
    """
    
    # Nombres alternativos permitidos para headers (case-insensitive, sin tildes)
    NOMBRES_HEADERS = [
        'NOMBRES COMPLETOS',
        'NOMBRE COMPLETO',
        'NOMBRES',
        'NOMBRE',
        'ESTUDIANTE',
        'ESTUDIANTES',
        'PARTICIPANTE',
        'PARTICIPANTES'
    ]
    
    CORREO_HEADERS = [
        'CORREO ELECTRONICO',
        'CORREO ELECTRÓNICO',
        'CORREOS ELECTRONICOS',
        'CORREOS ELECTRÓNICOS',
        'CORREO',
        'CORREOS',
        'EMAIL',
        'EMAILS',
        'E-MAIL',
        'E-MAILS',
        'MAIL',
        'MAILS'
    ]
    
    @staticmethod
    def normalize_text(text):
        """
        Normaliza texto para comparación: quita tildes, convierte a mayúsculas, quita espacios extras.
        """
        import unicodedata
        # Quitar tildes
        text = ''.join(
            c for c in unicodedata.normalize('NFD', str(text))
            if unicodedata.category(c) != 'Mn'
        )
        # Mayúsculas y quitar espacios extras
        return ' '.join(text.upper().split())
    
    def __init__(self, file_path_or_file):
        """
        Inicializa el parser con un archivo Excel.
        
        Args:
            file_path_or_file: Ruta al archivo  o File object de Django
        """
        self.file = file_path_or_file
        self.workbook = None
        self.worksheet = None
        self.nombres_col_index = None
        self.correo_col_index = None
    
    def parse(self) -> List[Dict[str, str]]:
        """
        Parsea el archivo Excel y retorna lista de estudiantes.
        
        Returns:
            Lista de diccionarios:
            [
                {"nombres_completos": "Juan Pérez", "correo_electronico": "juan@example.com"},
                ...
            ]
        
        Raises:
            ExcelParseError: Si el formato es inválido o hay errores
        """
        try:
            # Cargar workbook
            self.workbook = openpyxl.load_workbook(self.file, read_only=True, data_only=True)
            self.worksheet = self.workbook.active
            
            # Identificar columnas
            self._identify_columns()
            
            # Extraer datos
            estudiantes = self._extract_data()
            
            # Validar datos
            self._validate_data(estudiantes)
            
            logger.info(f"Excel parseado exitosamente: {len(estudiantes)} estudiante(s) encontrado(s)")
            
            return estudiantes
            
        except openpyxl.utils.exceptions.InvalidFileException:
            raise ExcelParseError("El archivo no es un Excel válido (.xlsx, .xls)")
        except Exception as e:
            logger.error(f"Error al parsear Excel: {str(e)}")
            raise ExcelParseError(f"Error al procesar el archivo: {str(e)}")
        finally:
            if self.workbook:
                self.workbook.close()
    
    def _identify_columns(self):
        """
        Identifica las columnas de nombres y correo en el header.
        
        Raises:
            ExcelParseError: Si no se encuentran las columnas requeridas
        """
        if not self.worksheet or self.worksheet.max_row < 1:
            raise ExcelParseError("El archivo Excel está vacío")
        
        # Leer headers (primera fila)
        headers = []
        for cell in self.worksheet[1]:
            value = str(cell.value).strip() if cell.value else ''
            headers.append(value)
        
        # Normalizar headers para comparación
        normalized_headers = [self.normalize_text(h) for h in headers]
        
        # Buscar columna de nombres
        for idx, norm_header in enumerate(normalized_headers):
            for nombre_option in self.NOMBRES_HEADERS:
                if self.normalize_text(nombre_option) in norm_header or norm_header in self.normalize_text(nombre_option):
                    self.nombres_col_index = idx
                    break
            if self.nombres_col_index is not None:
                break
        
        # Buscar columna de correo
        for idx, norm_header in enumerate(normalized_headers):
            for correo_option in self.CORREO_HEADERS:
                if self.normalize_text(correo_option) in norm_header or norm_header in self.normalize_text(correo_option):
                    self.correo_col_index = idx
                    break
            if self.correo_col_index is not None:
                break
        
        # Validar que se encontraron ambas columnas
        if self.nombres_col_index is None:
            raise ExcelParseError(
                f"No se encontró la columna de NOMBRES. "
                f"Headers encontrados: {', '.join(headers)}. "
                f"Use alguna de estas variantes: NOMBRES COMPLETOS, NOMBRE, ESTUDIANTE, PARTICIPANTE"
            )
        
        if self.correo_col_index is None:
            raise ExcelParseError(
                f"No se encontró la columna de CORREO ELECTRONICO. "
                f"Headers encontrados: {', '.join(headers)}. "
                f"Use alguna de estas variantes: CORREO ELECTRONICO, CORREO, EMAIL, MAIL"
            )
        
        logger.info(f"Columnas identificadas - Nombres: '{headers[self.nombres_col_index]}' (col {self.nombres_col_index}), "
                   f"Correo: '{headers[self.correo_col_index]}' (col {self.correo_col_index})")
    
    def _extract_data(self) -> List[Dict[str, str]]:
        """
        Extrae los datos de las columnas identificadas.
        
        Returns:
            Lista de diccionarios con datos crudos
        """
        estudiantes = []
        
        # Iterar desde la fila 2 (después del header) hasta el final
        for row_idx, row in enumerate(self.worksheet.iter_rows(min_row=2), start=2):
            nombres_cell = row[self.nombres_col_index]
            correo_cell = row[self.correo_col_index]
            
            nombres = str(nombres_cell.value).strip() if nombres_cell.value else ''
            correo = str(correo_cell.value).strip() if correo_cell.value else ''
            
            # Saltar filas vacías
            if not nombres and not correo:
                continue
            
            # Validar que ambos campos están presentes
            if not nombres:
                logger.warning(f"Fila {row_idx}: Nombre vacío, correo: {correo}")
                continue
            
            if not correo:
                logger.warning(f"Fila {row_idx}: Correo vacío, nombre: {nombres}")
                continue
            
            estudiantes.append({
                'nombres_completos': nombres,
                'correo_electronico': correo,
                'row_number': row_idx
            })
        
        if not estudiantes:
            raise ExcelParseError(
                "No se encontraron estudiantes en el archivo. "
                "Asegúrese de que hay datos después de la fila de encabezados."
            )
        
        return estudiantes
    
    def _validate_data(self, estudiantes: List[Dict[str, str]]):
        """
        Valida los datos extraídos.
        
        Args:
            estudiantes: Lista de estudiantes a validar
        
        Raises:
            ExcelParseError: Si hay errores de validación
        """
        errores = []
        correos_vistos = set()
        
        # Regex para validar email (solo permite letras, números, ., _, -, @ y dominios válidos)
        # No permite caracteres especiales raros
        email_regex = re.compile(r'^[a-zA-Z0-9][a-zA-Z0-9._-]*@[a-zA-Z0-9][a-zA-Z0-9.-]+\.[a-zA-Z]{2,}$')
        
        for estudiante in estudiantes:
            row_num = estudiante['row_number']
            nombres = estudiante['nombres_completos']
            correo = estudiante['correo_electronico']
            
            # Validar longitud de nombres
            if len(nombres) > 300:
                errores.append(f"Fila {row_num}: Nombre demasiado largo ({len(nombres)} caracteres, máximo 300)")
            
            # Validar que el nombre no esté vacío o sea solo espacios
            if not nombres or nombres.isspace():
                errores.append(f"Fila {row_num}: Nombre vacío o inválido")
            
            # Validar formato de correo
            correo_stripped = correo.strip()
            if not email_regex.match(correo_stripped):
                errores.append(
                    f"Fila {row_num}: Formato de correo inválido: {correo}. "
                    f"El correo solo puede contener letras, números, puntos, guiones y guión bajo."
                )
                continue
            
            # Validar que no tenga caracteres especiales raros (doble check)
            caracteres_invalidos = set(correo_stripped) - set('abcdefghijklmnopqrstuvwxyzABCDEFGHIJKLMNOPQRSTUVWXYZ0123456789@._-')
            if caracteres_invalidos:
                errores.append(
                    f"Fila {row_num}: Correo contiene caracteres no permitidos: {correo}. "
                    f"Caracteres inválidos: {', '.join(caracteres_invalidos)}"
                )
                continue
            
            # Validar que no empiece o termine con punto o guión
            local_part = correo_stripped.split('@')[0]
            if local_part.startswith('.') or local_part.startswith('-') or local_part.endswith('.') or local_part.endswith('-'):
                errores.append(f"Fila {row_num}: El correo no puede empezar o terminar con punto o guión: {correo}")
                continue
            
            # Validar que no tenga puntos consecutivos
            if '..' in correo_stripped:
                errores.append(f"Fila {row_num}: El correo no puede tener puntos consecutivos: {correo}")
                continue
            
            # Validar correos duplicados
            correo_lower = correo_stripped.lower()
            if correo_lower in correos_vistos:
                errores.append(f"Fila {row_num}: Correo duplicado: {correo}")
            else:
                correos_vistos.add(correo_lower)
        
        # Si hay errores, lanzar excepción con todos los detalles
        if errores:
            mensaje_error = "Errores de validación en el archivo Excel:\n" + "\n".join(errores)
            logger.error(mensaje_error)
            raise ExcelParseError(mensaje_error)


def parse_excel_estudiantes(file) -> List[Dict[str, str]]:
    """
    Función helper para parsear Excel de estudiantes.
    
    Args:
        file: Archivo de Django (UploadedFile)
    
    Returns:
        Lista de diccionarios con estudiantes validados
    
    Raises:
        ExcelParseError: Si hay errores de parsing o validación
    
    Ejemplo:
        >>> estudiantes = parse_excel_estudiantes(request.FILES['archivo_excel'])
        >>> for est in estudiantes:
        >>>     print(est['nombres_completos'], est['correo_electronico'])
    """
    parser = ExcelParser(file)
    return parser.parse()
